from openpyxl import load_workbook
wb = load_workbook('../material/11月考勤统计.xlsx')
ws = wb.active
info_dict = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    staff_id = row[0]
    staff_late = row[-1]
    info_dict[staff_id] = staff_late
monthly_wb = load_workbook('../material/迟到次数月度统计（11月更新）.xlsx')
monthly_ws = monthly_wb.active
for monthly_row in monthly_ws.iter_rows(min_row=3, max_col=14, values_only=True):
    member_id = monthly_row[0]
    member_late = monthly_row[-1]
    if member_late != info_dict[member_id]:
        print('工号{}迟到情况不匹配，请核查后更新'.format(member_id))