from openpyxl import load_workbook, Workbook
wb1 = load_workbook('../material/practice0.xlsx')
print(wb1)
new_wb = Workbook()
print(new_wb)
new_wb.save('../material/practice0_result.xlsx')
