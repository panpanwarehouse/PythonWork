from openpyxl import load_workbook, Workbook
wb = load_workbook('../material/10月考勤统计.xlsx')
ws = wb.active
late_header = []
for cell in ws[1]:
    late_header.append(cell.value)
new_wb = Workbook()
new_ws = new_wb.active
new_ws.append(late_header)
for row in ws.iter_rows(min_row = 2, values_only = True):
    name = row[1]
    time = row[3]
    number = row[-1]
    if time > 45 and number > 3:
        print('{}迟到了{}分钟，迟到了{}次'.format(name, time, number))
        new_ws.append(row)
new_wb.save('../material/人力资源部10月迟到人员信息.xlsx')