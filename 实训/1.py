from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('../material/practice1.xlsx')
# 获取活动工作表
active_ws = staff_wb.active

#然后打印出第5到第10行，前三列的所有数据（值）
for row in active_ws.iter_rows(min_row=5, max_row=10, min_col=0, max_col=3,values_only=True):
    print(row)

info_list = ['S1911', '萧爵瑟', 3000, '内容']

active_ws.append(info_list)

# 保存工作簿为【append_demo.xlsx】
staff_wb.save('../material/practice1_result.xlsx')
